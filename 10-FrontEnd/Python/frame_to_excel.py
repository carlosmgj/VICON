"""
frame_to_excel.py — Lee frame.raw y lo exporta a Excel con una fila por línea de imagen.
Uso: py -3.11 frame_to_excel.py [frame.raw] [W] [H]
"""

import sys
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

RAW_FILE = sys.argv[1] if len(sys.argv) > 1 else 'frame.raw'
W        = int(sys.argv[2]) if len(sys.argv) > 2 else 640
H        = int(sys.argv[3]) if len(sys.argv) > 3 else 480
OUT_FILE = RAW_FILE.replace('.raw', '.xlsx')

# Leer datos
with open(RAW_FILE, 'rb') as f:
    data = f.read()

print(f"Bytes leídos: {len(data)} (esperado {W*H})")

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Frame"

# Cabecera de columnas (índice de píxel)
header_fill = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
header_font = Font(color="FFFFFF", bold=True, size=8)

ws.cell(row=1, column=1, value="Fila\\Col")
ws.cell(row=1, column=1).fill = header_fill
ws.cell(row=1, column=1).font = header_font

for col in range(W):
    c = ws.cell(row=1, column=col + 2, value=col)
    c.fill = header_fill
    c.font = header_font
    c.alignment = Alignment(horizontal='center')
    ws.column_dimensions[get_column_letter(col + 2)].width = 4

ws.column_dimensions['A'].width = 8

# Datos — una fila Excel por línea de imagen
# Colorear cada celda en escala de grises según el valor
for row in range(min(H, len(data) // W)):
    # Etiqueta de fila
    r_cell = ws.cell(row=row + 2, column=1, value=row)
    r_cell.fill = header_fill
    r_cell.font = header_font
    r_cell.alignment = Alignment(horizontal='center')

    for col in range(W):
        idx = row * W + col
        if idx >= len(data):
            break
        val = data[idx]

        c = ws.cell(row=row + 2, column=col + 2, value=val)
        c.alignment = Alignment(horizontal='center')
        c.font = Font(size=7)

        # Color de fondo en escala de grises
        grey = format(val, '02X')
        c.fill = PatternFill(start_color=f"{grey}{grey}{grey}",
                             end_color=f"{grey}{grey}{grey}",
                             fill_type="solid")
        # Texto blanco si fondo oscuro
        if val < 128:
            c.font = Font(color="FFFFFF", size=7)

ws.freeze_panes = "B2"

wb.save(OUT_FILE)
print(f"Guardado en {OUT_FILE}")
